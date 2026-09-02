#!/usr/bin/env python3
"""Write one vendor QA workbook under output/, from the vendor template."""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import SITE_LABEL, ensure_in_product_output
from gate import assert_can_deliver, product_dir_from_path

ROOT = Path(__file__).resolve().parents[1]
HEADERS = {
    "A": "序号",
    "B": "站点",
    "C": "Asin（必填）",
    "D": "操作时间",
    "E": "提问内容",
    "F": "回答内容",
    "G": "备注",
}
THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def load_items(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return data
    items = data.get("items")
    if not isinstance(items, list):
        raise SystemExit("JSON must be a list or {items: [...]}")
    return items


def apply_cell(cell, value, *, center: bool, header_font: Font | None):
    cell.value = value
    cell.alignment = CENTER if center else WRAP
    cell.border = THIN
    if header_font:
        cell.font = Font(
            name=header_font.name or "宋体",
            size=11,
            bold=False,
            color="000000",
        )


def row_height(question: str, answer: str) -> float:
    chars = max(len(question), len(answer), 1)
    lines = max(2, min(8, (chars // 28) + 1))
    return 18.0 * lines


def clear_data_rows(ws) -> None:
    last = max(ws.max_row or 2, 2)
    for r in range(2, last + 1):
        for col in range(1, 8):
            ws.cell(r, col).value = None


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="QA JSON path")
    p.add_argument("--output", required=True, help="Must be output/{产品名}/{batch}_qa.xlsx")
    p.add_argument(
        "--template",
        default=str(ROOT / "templates" / "点赞、QA.xlsx"),
        help="Vendor workbook template (J1 instructions preserved)",
    )
    p.add_argument("--sheet", default="QA")
    p.add_argument(
        "--gate-ok",
        action="store_true",
        help="Internal: allow only after run_autoqa.py deliver validated",
    )
    args = p.parse_args()

    items = load_items(Path(args.input))
    if not items:
        print("no items")
        return 1

    out = ensure_in_product_output(Path(args.output), ROOT / "output")
    qa_json = Path(args.input).resolve()
    if not args.gate_ok:
        raise SystemExit(
            "Direct write blocked. Use:\n"
            f"  python scripts/run_autoqa.py deliver --qa-json {qa_json}"
        )
    product_dir = product_dir_from_path(qa_json, ROOT / "output")
    assert_can_deliver(product_dir, qa_json)

    template = Path(args.template)
    if not template.is_file():
        raise SystemExit(f"template not found: {template}")

    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, out)

    wb = load_workbook(out)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"missing sheet {args.sheet!r}; have {wb.sheetnames}")
    ws = wb[args.sheet]

    for col, expected in HEADERS.items():
        got = ws[f"{col}1"].value
        if got != expected:
            print(f"WARN header {col}1={got!r} expected {expected!r}", file=sys.stderr)

    header_font = copy(ws["A1"].font)
    clear_data_rows(ws)
    grouped = sorted(
        items,
        key=lambda x: (
            str(x.get("marketplace") or ""),
            str(x.get("asin") or "").upper(),
        ),
    )

    r = 2
    for seq, row in enumerate(grouped, 1):
        asin = str(row.get("asin") or "").strip().upper()
        mp = str(row.get("marketplace") or "").strip().upper()
        site = row.get("site") or SITE_LABEL.get(mp, mp)
        q = str(row["question"]).strip()
        a = str(row["answer"]).strip()
        note = str(row.get("note") or "").strip()
        values = {1: seq, 2: site, 3: asin, 4: None, 5: q, 6: a, 7: note or None}
        for col, val in values.items():
            apply_cell(
                ws.cell(r, col),
                val,
                center=col in (1, 2, 3, 4, 7),
                header_font=header_font,
            )
        ws.row_dimensions[r].height = row_height(q, a)
        r += 1

    for col in range(1, 8):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 16

    wb.save(out)
    print(f"wrote {len(grouped)} rows to {out} (vendor sheet only)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
